from openpyxl import load_workbook
import openpyxl

def Analysis(name_file: str, name_new_file: str):
    """
    Анализирует EXEL-файл магазина Мегабит и возращает EXEL-файл с лучшими скидками по алмазной карте по категориям
    """
    print('Подождите, анализ займет несколько секунд')
    columns = {
        1: 'Товар',
        2: 'Партномер',
        3: 'Артикул',
        4: 'Цена розница',
        8: 'Цена алмазная карта',
        9: 'Скидка'

    }
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = 'BestPrice'
    for i in range(1, 7):
        ws_new.cell(row=1, column=i).value = list(columns.values())[i-1]
    wb = load_workbook(name_file)
    sheet = wb['TDSheet']
    max_disc = 0
    max = []
    rows = 2
    for i in range(6, sheet.max_row):
        name = sheet.cell(row=i, column=1).value
        part_numb = sheet.cell(row=i, column=2).value
        article = sheet.cell(row=i, column=3).value
        price = sheet.cell(row=i, column=4).value
        disc_price = sheet.cell(row=i, column=8).value
        if disc_price:
            disc = round((price - disc_price)/price * 100)
            if disc > max_disc:
                max = [name, part_numb, article, price, disc_price, disc]
                max_disc = disc

        else:
            if max:
                ws_new.cell(row=rows, column=1).value, ws_new.cell(row=rows, column=2).value, ws_new.cell(row=rows, column=3).value,\
                ws_new.cell(row=rows, column=4).value, ws_new.cell(row=rows, column=5).value, ws_new.cell(row=rows, column=6).value = max
                rows += 1
            ws_new.cell(row=rows, column=1).value = name
            max_disc = 0
            max = []
            rows += 1
    wb_new.save(name_new_file)
    print('Файл сохранен в корневую папку проекта. Спасибо, что Вы с нами')


if __name__ == "__main__":
    Analysis('MegabitPrice.xlsx', 'BestPricesMegabit.xlsx')
